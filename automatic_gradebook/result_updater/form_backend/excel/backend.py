from ..__protocol import FormBackendProtocol
import openpyxl


class ExcelBackend(FormBackendProtocol):
    def __init__(self):
        super().__init__(None)

    def open(self, path):
        self._filename = path
        self._form = openpyxl.load_workbook(filename=path)
        self._const_form = openpyxl.load_workbook(filename=path,
                                                  data_only=True
                                                  )
        self._sheet = None
        self._val_sheet = None

    def save(self):
        from datetime import datetime
        date = datetime.date(datetime.now())
        filename = str(self._filename)  # "/".join(self._filename.split(".")[:-1]) + "_" + str(date) + ".xlsx"
        self._const_form.close()
        self._form.save(filename)

    def create(self, path):
        wb = openpyxl.Workbook()
        wb.save(path)
        wb.close()
        del wb
        self.open(path)

    def select_sheet(self, sheet_name):
        self._sheet = self._form[sheet_name]
        self._val_sheet = self._const_form[sheet_name]

    def get_cell_value(self, cell_addr):
        return self._val_sheet[cell_addr].value

    def set_cell_value(self, cell_addr, value):
        self._sheet[cell_addr] = value
        self._val_sheet[cell_addr] = value

    def set_cell_color(self, cell_addr, color):
        c = openpyxl.styles.colors.Color(rgb=color)
        self._sheet[cell_addr].fill = openpyxl.styles.fills.PatternFill(patternType='solid',
                                                                        start_color=c)

    def insert_row(self, ind):
        self._sheet.insert_row(ind)
        self._val_sheet.insert_row(ind)
